"""
Schedule generator for work rotation patterns.
Handles multiple employees with different rotation states and shift assignments.
"""

import csv
import os
from datetime import date, timedelta
from typing import Dict, List, Tuple
from employee import Employee
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import calendar

class ScheduleGenerator:
    """
    Generates work schedules based on rotation patterns and shift assignments.
    """
    
    def __init__(self, start_date: date):
        """
        Initialize schedule generator.
        
        Args:
            start_date: Starting date for schedule generation
        """
        self.start_date = start_date
        # Reference date for continuity (August 1, 2025)
        self.reference_date = date(2025, 8, 1)
        self.employees = {
            'Ludy': Employee.create_ludy(),
            'Isaac': Employee.create_isaac(),
            'Genesis': Employee.create_genesis()
        }
        # Calculate days from reference and advance employee states
        self._advance_to_start_date()
    
    def _advance_to_start_date(self):
        """
        Advance employee states to match the start date based on August 1, 2025 reference.
        """
        if self.start_date >= self.reference_date:
            days_to_advance = (self.start_date - self.reference_date).days
            
            # Advance each employee's state
            for employee in self.employees.values():
                for _ in range(days_to_advance):
                    employee.advance_day()
    
    def validate_daily_schedule(self, day_schedule: Dict[str, dict]) -> bool:
        """
        Validate that exactly 2 employees are working and 1 is resting.
        
        Args:
            day_schedule: Dictionary with employee schedules for a day
            
        Returns:
            True if schedule is valid, False otherwise
        """
        working_count = sum(1 for emp_data in day_schedule.values() 
                          if emp_data['status'] == 'Trabajando')
        return working_count == 2
    
    def generate_schedule(self, num_days: int) -> Tuple[List[Dict], Dict[str, List[Dict]]]:
        """
        Generate complete schedule for all employees.
        
        Args:
            num_days: Number of days to generate
            
        Returns:
            Tuple of (general_schedule, individual_schedules)
        """
        general_schedule = []
        individual_schedules = {name: [] for name in self.employees.keys()}
        
        current_date = self.start_date
        
        for day in range(num_days):
            day_schedule = {}
            
            # Get schedule for each employee for current day
            for name, employee in self.employees.items():
                schedule_info = employee.get_schedule_info()
                day_schedule[name] = schedule_info
                
                # Add to individual schedule with date
                individual_record = {
                    'date': current_date.strftime('%Y-%m-%d'),
                    'weekday': current_date.strftime('%A'),
                    **schedule_info
                }
                individual_schedules[name].append(individual_record)
            
            # Validate schedule (exactly 2 working, 1 resting)
            if not self.validate_daily_schedule(day_schedule):
                raise ValueError(f"Invalid schedule on {current_date}: "
                               f"Must have exactly 2 employees working and 1 resting")
            
            # Add date information to general schedule (translate weekday to Spanish)
            spanish_weekdays = {
                'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
                'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
            }
            weekday_spanish = spanish_weekdays.get(current_date.strftime('%A'), current_date.strftime('%A'))
            
            general_record = {
                'Fecha': current_date.strftime('%Y-%m-%d'),
                'Día': weekday_spanish,
                **day_schedule
            }
            general_schedule.append(general_record)
            
            # Advance all employees to next day
            for employee in self.employees.values():
                employee.advance_day()
            
            current_date += timedelta(days=1)
        
        return general_schedule, individual_schedules
    
    def create_calendar_format(self, general_schedule: List[Dict], start_date: date) -> Dict[str, List[List]]:
        """
        Create monthly calendar format for Excel.
        
        Args:
            general_schedule: General schedule data
            start_date: Starting date for schedule
            
        Returns:
            Dictionary with monthly calendar data
        """
        # Spanish month and day names
        spanish_months = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]
        spanish_days = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
        
        monthly_calendars = {}
        current_date = start_date
        schedule_index = 0
        
        # Group schedules by month
        while schedule_index < len(general_schedule):
            month_key = f"{spanish_months[current_date.month - 1]} {current_date.year}"
            
            # Get calendar structure for this month
            cal = calendar.monthcalendar(current_date.year, current_date.month)
            
            # Create calendar grid with headers
            calendar_grid = []
            
            # Add day headers
            calendar_grid.append(spanish_days)
            
            # Fill calendar with schedule data
            for week in cal:
                week_data = []
                for day in week:
                    if day == 0:
                        week_data.append("")  # Empty cell for days not in this month
                    else:
                        day_date = date(current_date.year, current_date.month, day)
                        
                        # Find corresponding schedule data
                        if (day_date >= start_date and 
                            schedule_index < len(general_schedule) and
                            day_date == start_date + timedelta(days=schedule_index)):
                            
                            day_schedule = general_schedule[schedule_index]
                            
                            # Create cell content with employee info and color info
                            cell_content = []
                            cell_colors = []
                            for emp_name in ['Ludy', 'Isaac', 'Genesis']:
                                emp_data = day_schedule[emp_name]
                                if emp_data['status'] == 'Trabajando':
                                    cell_content.append(f"{emp_name}: {emp_data['shift']}")
                                    if emp_data['shift'] == 'Mañana':
                                        cell_colors.append('morning')
                                    else:
                                        cell_colors.append('afternoon')
                                else:
                                    cell_content.append(f"{emp_name}: Desc")
                                    cell_colors.append('rest')
                            
                            week_data.append({
                                'day': day,
                                'content': cell_content,
                                'colors': cell_colors,
                                'schedule': day_schedule
                            })
                            schedule_index += 1
                        else:
                            week_data.append({'day': day, 'content': [], 'schedule': None})
                
                calendar_grid.append(week_data)
            
            monthly_calendars[month_key] = calendar_grid
            
            # Move to next month
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
            
            # Stop if we've processed all schedule data
            if schedule_index >= len(general_schedule):
                break
        
        return monthly_calendars
    
    def save_to_excel(self, general_schedule: List[Dict], 
                     individual_schedules: Dict[str, List[Dict]], 
                     output_dir: str):
        """
        Save schedules to Excel file with multiple worksheets.
        
        Args:
            general_schedule: General schedule data
            individual_schedules: Individual employee schedules
            output_dir: Output directory for Excel file
        """
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        # Create workbook
        wb = Workbook()
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        day_header_font = Font(bold=True, size=12)
        
        # Colors for different statuses (updated palette)
        morning_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Azul claro
        afternoon_fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")  # Naranja claro
        rest_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
        
        # Darker versions for name labels
        morning_dark = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")  # Azul
        afternoon_dark = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")  # Naranja
        rest_dark = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")  # Verde
        
        # Create monthly calendar format
        monthly_calendars = self.create_calendar_format(general_schedule, self.start_date)
        
        # Spanish weekday names
        spanish_weekdays = {
            'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
            'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
        }
        
        # Create monthly calendar worksheets with improved structure
        for month_name, calendar_data in monthly_calendars.items():
            cal_ws = wb.create_sheet(title=f"Calendario {month_name}")
            
            # Month title
            cal_ws.merge_cells('A1:G1')
            title_cell = cal_ws.cell(row=1, column=1, value=month_name)
            title_cell.font = Font(bold=True, size=16, color="FFFFFF")
            title_cell.alignment = center_align
            title_cell.fill = header_fill
            
            # Day headers (row 2)
            for col, day_name in enumerate(calendar_data[0], 1):
                cell = cal_ws.cell(row=2, column=col, value=day_name)
                cell.font = day_header_font
                cell.alignment = center_align
                cell.border = border
            
            # Calendar days with improved structure (4 rows per week)
            base_row = 3
            for week_idx, week_data in enumerate(calendar_data[1:]):
                # Calculate starting row for this week (4 rows per week: day number + 3 employees)
                week_start_row = base_row + (week_idx * 4)
                
                # Set row heights for the week structure
                cal_ws.row_dimensions[week_start_row].height = 25  # Day number row
                cal_ws.row_dimensions[week_start_row + 1].height = 30  # Ludy row  
                cal_ws.row_dimensions[week_start_row + 2].height = 30  # Isaac row
                cal_ws.row_dimensions[week_start_row + 3].height = 30  # Genesis row
                
                for col, day_data in enumerate(week_data, 1):
                    if day_data == "":  # Empty cell for days not in this month
                        continue
                    elif isinstance(day_data, dict) and day_data['schedule']:
                        # Day number cell
                        day_cell = cal_ws.cell(row=week_start_row, column=col, value=str(day_data['day']))
                        day_cell.alignment = center_align
                        day_cell.border = border
                        day_cell.font = Font(size=11, bold=True)
                        day_cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                        
                        # Employee cells with individual colors
                        employee_names = ['Ludy', 'Isaac', 'Genesis']
                        for i, emp_name in enumerate(employee_names):
                            emp_data = day_data['schedule'][emp_name]
                            emp_row = week_start_row + 1 + i
                            emp_cell = cal_ws.cell(row=emp_row, column=col)
                            
                            if emp_data['status'] == 'Trabajando':
                                emp_cell.value = f"{emp_name}: {emp_data['shift']}"
                                if emp_data['shift'] == 'Mañana':
                                    emp_cell.fill = morning_fill
                                else:
                                    emp_cell.fill = afternoon_fill
                            else:
                                emp_cell.value = f"{emp_name}: Descanso"
                                emp_cell.fill = rest_fill
                            
                            emp_cell.alignment = center_align
                            emp_cell.border = border
                            emp_cell.font = Font(size=8, bold=True)
                    
                    elif isinstance(day_data, dict):
                        # Day without schedule data
                        day_cell = cal_ws.cell(row=week_start_row, column=col, value=str(day_data['day']))
                        day_cell.alignment = center_align
                        day_cell.border = border
            
            # Add color legend below calendar
            legend_start_row = base_row + (len(calendar_data[1:]) * 4) + 2
            cal_ws.cell(row=legend_start_row, column=1, value="LEYENDA:").font = Font(bold=True, size=12)
            
            # Morning legend
            morning_legend = cal_ws.cell(row=legend_start_row + 1, column=1, value="Mañana")
            morning_legend.fill = morning_fill
            morning_legend.font = Font(bold=True)
            morning_legend.border = border
            morning_legend.alignment = center_align
            
            # Afternoon legend  
            afternoon_legend = cal_ws.cell(row=legend_start_row + 1, column=2, value="Tarde")
            afternoon_legend.fill = afternoon_fill
            afternoon_legend.font = Font(bold=True)
            afternoon_legend.border = border
            afternoon_legend.alignment = center_align
            
            # Rest legend
            rest_legend = cal_ws.cell(row=legend_start_row + 1, column=3, value="Descanso")
            rest_legend.fill = rest_fill
            rest_legend.font = Font(bold=True)
            rest_legend.border = border
            rest_legend.alignment = center_align
            
            # Auto-adjust column widths for calendar
            for col in range(1, 8):
                column_letter = get_column_letter(col)
                cal_ws.column_dimensions[column_letter].width = 18
        
        # Create general schedule worksheet
        general_ws = wb.active
        if "Calendario" in wb.sheetnames[0]:
            general_ws = wb.create_sheet(title="Lista General", index=0)
            wb.active = general_ws
        else:
            general_ws.title = "Lista General"
        
        # Headers for general schedule
        headers = ['Fecha', 'Día', 'Ludy', 'Isaac', 'Génesis']
        for col, header in enumerate(headers, 1):
            cell = general_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Data rows for general schedule
        current_date = self.start_date
        for row, day_schedule in enumerate(general_schedule, 2):
            # Date and weekday
            general_ws.cell(row=row, column=1, value=current_date.strftime('%Y-%m-%d')).border = border
            weekday_spanish = spanish_weekdays[current_date.strftime('%A')]
            general_ws.cell(row=row, column=2, value=weekday_spanish).border = border
            
            # Employee statuses
            col = 3
            for emp_name in ['Ludy', 'Isaac', 'Genesis']:
                emp_data = day_schedule[emp_name]
                if emp_data['status'] == 'Trabajando':
                    cell_value = f"{emp_data['shift']}"
                    cell = general_ws.cell(row=row, column=col, value=cell_value)
                    if emp_data['shift'] == 'Mañana':
                        cell.fill = morning_fill
                    else:
                        cell.fill = afternoon_fill
                else:
                    cell = general_ws.cell(row=row, column=col, value="Descanso")
                    cell.fill = rest_fill
                
                cell.border = border
                cell.alignment = center_align
                col += 1
            
            current_date += timedelta(days=1)
        
        # Auto-adjust column widths for general schedule
        for col in range(1, 6):
            column_letter = get_column_letter(col)
            max_length = max(len(str(general_ws.cell(row=row, column=col).value or "")) 
                           for row in range(1, len(general_schedule) + 2))
            general_ws.column_dimensions[column_letter].width = min(max_length + 2, 20)
        
        # Create individual employee calendars
        for emp_name in ['Ludy', 'Isaac', 'Genesis']:
            # Individual calendar view
            cal_ws = wb.create_sheet(title=f"Cal. {emp_name}")
            
            # Create individual monthly calendars for this employee
            current_date = self.start_date
            schedule_data = individual_schedules[emp_name]
            row_offset = 1
            
            for month_name, calendar_data in monthly_calendars.items():
                # Month title
                cal_ws.merge_cells(f'A{row_offset}:G{row_offset}')
                title_cell = cal_ws.cell(row=row_offset, column=1, value=f"{emp_name} - {month_name}")
                title_cell.font = Font(bold=True, size=14, color="FFFFFF")
                title_cell.alignment = center_align
                title_cell.fill = header_fill
                row_offset += 1
                
                # Day headers
                for col, day_name in enumerate(calendar_data[0], 1):
                    cell = cal_ws.cell(row=row_offset, column=col, value=day_name)
                    cell.font = Font(bold=True, size=10)
                    cell.alignment = center_align
                    cell.border = border
                row_offset += 1
                
                # Calendar days for this employee
                for week_data in calendar_data[1:]:
                    cal_ws.row_dimensions[row_offset].height = 40
                    for col, day_data in enumerate(week_data, 1):
                        if day_data == "":  # Empty cell
                            continue
                        elif isinstance(day_data, dict) and day_data['schedule']:
                            emp_data = day_data['schedule'][emp_name]
                            
                            if emp_data['status'] == 'Trabajando':
                                if emp_data['shift'] == 'Mañana':
                                    cell_text = f"{day_data['day']}\n🔵 {emp_data['shift']}"
                                    cell_fill = morning_fill
                                else:
                                    cell_text = f"{day_data['day']}\n🟠 {emp_data['shift']}"
                                    cell_fill = afternoon_fill
                            else:
                                cell_text = f"{day_data['day']}\n🟢 Descanso"
                                cell_fill = rest_fill
                            
                            cell = cal_ws.cell(row=row_offset, column=col, value=cell_text)
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.border = border
                            cell.fill = cell_fill
                            cell.font = Font(size=9, bold=True)
                        elif isinstance(day_data, dict):
                            # Day without schedule data
                            cell = cal_ws.cell(row=row_offset, column=col, value=str(day_data['day']))
                            cell.alignment = center_align
                            cell.border = border
                    row_offset += 1
                
                # Add spacing between months
                row_offset += 2
            
            # Auto-adjust column widths for individual calendar
            for col in range(1, 8):
                column_letter = get_column_letter(col)
                cal_ws.column_dimensions[column_letter].width = 12
        
        # Create individual schedule worksheets (list format)
        for emp_name, schedule_data in individual_schedules.items():
            ws = wb.create_sheet(title=f"Lista {emp_name}")
            
            # Headers for individual schedule
            ind_headers = ['Fecha', 'Día', 'Estado', 'Turno', 'Ciclo', 'Día en Ciclo']
            for col, header in enumerate(ind_headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            
            # Data rows for individual schedule
            for row, day_data in enumerate(schedule_data, 2):
                ws.cell(row=row, column=1, value=day_data['date']).border = border
                # Convert weekday to Spanish
                weekday_english = day_data['weekday']
                weekday_spanish = spanish_weekdays.get(weekday_english, weekday_english)
                ws.cell(row=row, column=2, value=weekday_spanish).border = border
                
                # Status cell with color coding
                status_cell = ws.cell(row=row, column=3, value=day_data['status'])
                status_cell.border = border
                status_cell.alignment = center_align
                if day_data['status'] == 'Trabajando':
                    if day_data['shift'] == 'Mañana':
                        status_cell.fill = morning_fill
                    else:
                        status_cell.fill = afternoon_fill
                else:
                    status_cell.fill = rest_fill
                
                # Shift cell
                shift_cell = ws.cell(row=row, column=4, value=day_data['shift'] if day_data['shift'] != '-' else '')
                shift_cell.border = border
                shift_cell.alignment = center_align
                
                # Cycle and day in cycle
                ws.cell(row=row, column=5, value=day_data['cycle']).border = border
                ws.cell(row=row, column=6, value=day_data['day_in_cycle']).border = border
            
            # Auto-adjust column widths for individual schedule
            for col in range(1, 7):
                column_letter = get_column_letter(col)
                max_length = max(len(str(ws.cell(row=row, column=col).value or "")) 
                               for row in range(1, len(schedule_data) + 2))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 15)
        
        # Save the workbook
        excel_file = os.path.join(output_dir, 'calendario_trabajo.xlsx')
        wb.save(excel_file)
        return excel_file
    
    def save_to_csv(self, general_schedule: List[Dict], 
                   individual_schedules: Dict[str, List[Dict]], 
                   output_dir: str):
        """
        Save schedules to CSV files for Excel import.
        
        Args:
            general_schedule: General schedule data
            individual_schedules: Individual employee schedules
            output_dir: Output directory for CSV files
        """
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        # Save general schedule
        general_file = os.path.join(output_dir, 'general_schedule.csv')
        with open(general_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header
            header = ['Date', 'Weekday']
            for emp_name in self.employees.keys():
                header.extend([f'{emp_name}_Status', f'{emp_name}_Shift', f'{emp_name}_Cycle'])
            writer.writerow(header)
            
            # Data rows
            current_date = self.start_date
            for day_schedule in general_schedule:
                row = [current_date.strftime('%Y-%m-%d'), current_date.strftime('%A')]
                
                for emp_name in self.employees.keys():
                    emp_data = day_schedule[emp_name]
                    row.extend([
                        emp_data['status'],
                        emp_data['shift'],
                        emp_data['cycle']
                    ])
                
                writer.writerow(row)
                current_date += timedelta(days=1)
        
        # Save individual schedules
        for emp_name, schedule_data in individual_schedules.items():
            emp_file = os.path.join(output_dir, f'{emp_name.lower()}_schedule.csv')
            with open(emp_file, 'w', newline='', encoding='utf-8') as f:
                if schedule_data:
                    writer = csv.DictWriter(f, fieldnames=schedule_data[0].keys())
                    writer.writeheader()
                    writer.writerows(schedule_data)
    
    def print_schedule_summary(self, general_schedule: List[Dict], days_to_show: int = 14):
        """
        Print a summary of the schedule to console.
        
        Args:
            general_schedule: General schedule data
            days_to_show: Number of days to display
        """
        print(f"\nSchedule Summary (first {days_to_show} days):")
        print("=" * 100)
        print(f"{'Date':<12} {'Weekday':<10} {'Ludy':<20} {'Isaac':<20} {'Genesis':<20}")
        print("-" * 100)
        
        current_date = self.start_date
        for i, day_schedule in enumerate(general_schedule[:days_to_show]):
            date_str = current_date.strftime('%Y-%m-%d')
            weekday = current_date.strftime('%A')
            
            row = f"{date_str:<12} {weekday:<10} "
            
            for emp_name in ['Ludy', 'Isaac', 'Genesis']:
                emp_data = day_schedule[emp_name]
                status_shift = f"{emp_data['status']}"
                if emp_data['status'] == 'Working':
                    status_shift += f" ({emp_data['shift']})"
                row += f"{status_shift:<20} "
            
            print(row)
            current_date += timedelta(days=1)
        
        print("=" * 100)
